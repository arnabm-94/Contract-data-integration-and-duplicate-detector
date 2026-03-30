import pandas as pd
import numpy as np
from datetime import timedelta
import openpyxl
from openpyxl.styles import PatternFill

# Read the excel files
try:
    isg_file = 'ISG_File.xlsx'
    idc_file = 'IDC_File.xlsx'
    sector_mapping_file = 'CG Sector Mapping for Automation_Final.xlsx'

    df_isg = pd.read_excel(isg_file)
    df_idc = pd.read_excel(idc_file)
    df_sector_mapping = pd.read_excel(sector_mapping_file)
except Exception as e:
    print(f"Error reading files: {e}")
    raise

# Column mapping
column_mapping = {
    "Customer Name": "Company Name",
    "Customer Sub Industry": "Company Industry",
    "Vendor Name": "Service Providers",
    "Signing Country": "Country",
    "Services Contract Value (Base) (in US Dollar - USD)": "Tcv (Usd)",
    "Start Date": "Contract Start Date",
    "Contract Length (Base)": "Contract Length In Months",
    "End Date (Base)": "Renewal Date",
    "Deal Description": "Domains"
}

# Rename columns in ISG file to match IDC file columns
try:
    df_isg.rename(columns={v: k for k, v in column_mapping.items()}, inplace=True)
except Exception as e:
    print(f"Error in renaming columns: {e}")
    raise

# Select and order columns as per IDC file
try:
    df_isg = df_isg[list(column_mapping.keys())]
    df_idc = df_idc[list(column_mapping.keys())]
except KeyError as e:
    print(f"Error in selecting columns: {e}")
    raise

# Concatenate dataframes
df_combined = pd.concat([df_idc, df_isg], ignore_index=True)

# Convert Start Date to datetime
df_combined['Start Date'] = pd.to_datetime(df_combined['Start Date'], errors='coerce')

# Merge with CG Sector Mapping data based on "Customer Sub Industry" and "IDC Sub Vertical"
try:
    # First merge attempt: match "Customer Sub Industry" with "IDC Sub Vertical"
    df_combined = df_combined.merge(
        df_sector_mapping[['IDC Sub Vertical', 'CG Sector']],
        left_on='Customer Sub Industry',
        right_on='IDC Sub Vertical',
        how='left'
    )
    
    # Rename the columns to differentiate between the first and second merge attempts
    df_combined.rename(columns={'CG Sector': 'CG Sector IDC'}, inplace=True)

    # Second merge attempt: match "Customer Sub Industry" with "ISG Sub Vertical" where there was no match in the first attempt
    df_combined = df_combined.merge(
        df_sector_mapping[['ISG Sub Vertical', 'CG Sector']],
        left_on='Customer Sub Industry',
        right_on='ISG Sub Vertical',
        how='left'
    )
    
    # Fill the "CG Sector" column with values from the second merge where it is NaN in the first merge
    df_combined['CG Sector'] = df_combined['CG Sector IDC'].combine_first(df_combined['CG Sector'])
    
    # Drop the temporary columns used for merging
    df_combined.drop(columns=['CG Sector IDC'], inplace=True)

except Exception as e:
    print(f"Error merging with sector mapping: {e}")
    raise

# Sort the DataFrame by Vendor Name only
df_combined = df_combined.sort_values(by='Vendor Name').reset_index(drop=True)

# Function to highlight duplicates based on Vendor Name, Customer Name, and Start Date within 30 days range
def highlight_duplicates(df):
    duplicated_indices = []
    df = df.sort_values(by=['Vendor Name']).reset_index(drop=True)
    #df = df.sort_values(by=['Vendor Name', 'Customer Name', 'Start Date']).reset_index(drop=True)
    
    for i in range(1, len(df)):
        if (df.loc[i, 'Vendor Name'] == df.loc[i-1, 'Vendor Name'] and
            df.loc[i, 'Customer Name'] == df.loc[i-1, 'Customer Name'] and
            abs((df.loc[i, 'Start Date'] - df.loc[i-1, 'Start Date']).days) <= 30):
            duplicated_indices.append(i)
            duplicated_indices.append(i-1)
    
    return list(set(duplicated_indices))

# Highlight duplicates
duplicates = highlight_duplicates(df_combined)

# Save to a new excel file and highlight duplicates
try:
    with pd.ExcelWriter('highlighted_duplicates.xlsx', engine='openpyxl') as writer:
        df_combined.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Define the yellow fill for duplicates
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        for row in duplicates:
            for col in range(1, len(df_combined.columns) + 1):
                worksheet.cell(row=row + 2, column=col).fill = yellow_fill
except Exception as e:
    print(f"Error saving the Excel file: {e}")
    raise

print("The new Excel file 'contract_analysis_document.xlsx' has been created.")
